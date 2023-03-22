import sqlite3

from openpyxl import load_workbook
import pandas as pd


def create_table_database(exel_file, DB_name):
    con = sqlite3.connect('data.db')
    wb = pd.read_excel('example.xlsx', sheet_name=None)
    for sheet in wb:
        wb[sheet].to_sql(sheet, con, index=False)

        # так много одинаковых запросов на создание новых столбцов, потому что
        # SQLite не поддерживает добавление нескольких столбцов в таблицу с
        # помощью одного оператора
    con.execute("ALTER TABLE 'Лист1' ADD COLUMN Total_Qliq_data1 INTEGER")
    con.execute("ALTER TABLE 'Лист1' ADD COLUMN Total_Qliq_data2 INTEGER")
    con.execute("ALTER TABLE 'Лист1' ADD COLUMN Total_Qoil_data1 INTEGER")
    con.execute("ALTER TABLE 'Лист1' ADD COLUMN Total_Qoil_data2 INTEGER")
    con.commit()
    con.close()


def add_calculated_total():
    workbook = load_workbook(filename="example.xlsx", read_only=True)
    worksheet = workbook.active
    con = sqlite3.connect("data.db")

    # парсим excel-файл и записываем данные в базу данных
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        Total_Qliq_data1 = row[2] + row[6]
        Total_Qliq_data2 = row[3] + row[7]
        Total_Qoil_data1 = row[4] + row[8]
        Total_Qoil_data2 = row[5] + row[9]

        # записываем данные в базу данных
        con.execute("""
                    UPDATE Лист1 SET Total_Qliq_data1=?, Total_Qliq_data2=?, 
                    Total_Qoil_data1=?, Total_Qoil_data2=?
                    WHERE id=?;
                """, (Total_Qliq_data1, Total_Qliq_data2, Total_Qoil_data1,
                      Total_Qoil_data2, row[0]))

    # сохраняем изменения в базе данных
    con.commit()

    # закрываем соединение с базой данных
    con.close()


if __name__ == '__main__':
    try:
        create_table_database(exel_file='example.xlsx', DB_name='data.db')
    except Exception:
        print('БД с таким именем уже создана')
    finally:
        add_calculated_total()

